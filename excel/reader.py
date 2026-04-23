# -*- coding: utf-8 -*-
import openpyxl
from config import SHEET_NAME, DATA_START_ROW, INPUT_COLS, OUTPUT_COL_G, OUTPUT_COL_N


def read_requirements(filepath: str) -> list:
    """
    读取 Excel 文件中的需求行。
    返回列表，每项结构：
    {
        "row": int,          # 行号
        "context": str,      # 拼接的需求上下文（供 AI 使用）
        "module": str,       # 功能模块
        "feature": str,      # 需求名称
        "detail": str,       # 需求描述
        "extra": str,        # 补充说明
        "existing_g": str,   # G列已有内容
        "existing_n": float, # N列已有数值
    }
    """
    wb = openpyxl.load_workbook(filepath)

    # 支持的工作表名称列表（按优先级排序）
    supported_sheet_names = [
        SHEET_NAME,
        "需求拆解评估",
        "工时评估表AI版本",
        "工时评估表",
        "需求清单",
        "需求列表",
        "需求",
        "Sheet1",
    ]

    # 查找可用的工作表
    ws = None
    for name in supported_sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    
    # 如果没有找到预定义的工作表，使用第一个工作表
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    elif ws is None:
        raise ValueError("Excel 文件中没有找到任何工作表")

    max_row = ws.max_row

    # 处理合并单元格：向下填充 A 列（功能模块）
    module_map = _fill_down(ws, INPUT_COLS["module"], DATA_START_ROW, max_row)

    rows = []
    for r in range(DATA_START_ROW, max_row + 1):
        module  = module_map.get(r, "") or ""
        feature = _cell_val(ws, r, INPUT_COLS["feature"])
        detail  = _cell_val(ws, r, INPUT_COLS["detail"])
        extra   = _cell_val(ws, r, INPUT_COLS["extra"])

        # 跳过没有任何有效需求内容的行
        if not any([feature, detail, extra]):
            continue

        existing_g = _cell_val(ws, r, OUTPUT_COL_G)
        existing_n_raw = ws.cell(r, OUTPUT_COL_N).value
        try:
            existing_n = float(existing_n_raw) if existing_n_raw not in (None, "") else None
        except (ValueError, TypeError):
            existing_n = None

        # 拼接上下文
        parts = []
        if module:  parts.append(f"阶段/模块：{module}")
        if feature: parts.append(f"需求名称：{feature}")
        if detail:  parts.append(f"需求描述：{detail}")
        if extra:   parts.append(f"补充说明：{extra}")
        context = "  ".join(parts)

        rows.append({
            "row":        r,
            "context":    context,
            "module":     module,
            "feature":    feature,
            "detail":     detail,
            "extra":      extra,
            "existing_g": existing_g,
            "existing_n": existing_n,
        })

    return rows


def _fill_down(ws, col: int, start_row: int, end_row: int) -> dict:
    """追踪最近非空值，处理合并单元格向下填充。"""
    last_val = None
    result = {}
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, col).value
        if v is not None:
            last_val = str(v).strip()
        result[r] = last_val
    return result


def _cell_val(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""
