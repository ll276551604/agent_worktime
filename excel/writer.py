# -*- coding: utf-8 -*-
import os
import openpyxl
from datetime import datetime
from config import SHEET_NAME, OUTPUT_COL_G, OUTPUT_COL_N, OUTPUT_FOLDER


def write_results(src_filepath: str, results: list) -> str:
    """
    将 AI 拆解结果回填到 Excel 并保存副本。

    results 每项结构：
    {
        "row": int,
        "g_column_text": str,   # G列：页面×功能点文本 + 产品工时X天
        "days": float,          # N列：工时天数（数字，可求和）
        "skipped": bool,
    }

    返回：输出文件的绝对路径
    """
    wb = openpyxl.load_workbook(src_filepath)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"工作表 '{SHEET_NAME}' 不存在")

    ws = wb[SHEET_NAME]

    for item in results:
        if item.get("skipped"):
            continue

        row    = item["row"]
        g_text = item.get("g_column_text", "")
        days   = item.get("days")

        if g_text:
            ws.cell(row, OUTPUT_COL_G).value = g_text

        if days is not None:
            ws.cell(row, OUTPUT_COL_N).value = float(days)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    timestamp     = datetime.now().strftime("%Y%m%d_%H%M%S")
    original_name = os.path.basename(src_filepath)
    output_path   = os.path.join(OUTPUT_FOLDER, f"output_{timestamp}_{original_name}")
    wb.save(output_path)
    return output_path
