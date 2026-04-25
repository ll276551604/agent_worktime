# -*- coding: utf-8 -*-
import time
import sys
import os
import logging
import hashlib
from typing import List, Dict, Any

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logger = logging.getLogger(__name__)

from excel import reader, writer
from agent.graph import build_graph, KnowledgeLoader
from agent.knowledge_manager import get_knowledge_manager
from config import REQUEST_DELAY, AppConfig

# 评估结果缓存（简单的内存缓存）
_evaluation_cache = {}
_CACHE_MAX_SIZE = AppConfig.EVALUATION_CACHE_MAX_SIZE


def _generate_cache_key(req: Dict) -> str:
    """生成需求的唯一缓存键"""
    content = f"{req.get('module','')}|{req.get('feature','')}|{req.get('detail','')}"
    return hashlib.md5(content.encode()).hexdigest()


def _get_cached_result(req: Dict) -> Dict:
    """获取缓存的评估结果"""
    key = _generate_cache_key(req)
    return _evaluation_cache.get(key)


def _set_cached_result(req: Dict, result: Dict):
    """设置评估结果到缓存"""
    key = _generate_cache_key(req)
    
    # 淘汰策略：超过最大缓存数时删除最早的记录
    if len(_evaluation_cache) >= _CACHE_MAX_SIZE:
        oldest_key = next(iter(_evaluation_cache))
        del _evaluation_cache[oldest_key]
    
    _evaluation_cache[key] = result


def run_agent(filepath: str, skip_filled: bool, api_key: str = None,
              model_id: str = None, progress_callback=None, skill_id: str = None) -> str:
    """
    完整工时拆解流程：读取 → LangGraph 智能评估（含技能注入）→ 回填 Excel（G列+N列）

    progress_callback: fn(current, total, row_num, status, preview, page_count)
    """
    from agent import skill_manager as sm

    rows = reader.read_requirements(filepath)
    if not rows:
        raise ValueError("未找到任何需求行，请检查 Excel 格式和工作表名称")

    # 加载技能配置（整批共享）
    sid          = skill_id or sm.get_current_skill_id()
    skill_config = sm.get_skill(sid)
    examples     = sm.load_examples(sid, limit=3)

    kb    = KnowledgeLoader().load()
    graph = build_graph()

    total   = len(rows)
    results = []

    for i, row_data in enumerate(rows):
        row_num = row_data["row"]

        # 跳过已填写的行
        if skip_filled and row_data.get("existing_g"):
            results.append({
                "row":          row_num,
                "g_column_text": row_data["existing_g"],
                "days":         row_data["existing_n"],
                "skipped":      True,
            })
            if progress_callback:
                progress_callback(i + 1, total, row_num, "skipped", "（已有内容，跳过）", 0)
            continue

        print(f"[Agent] 处理行 {row_num} ({i+1}/{total}) skill={sid}", flush=True)
        try:
            req = {
                "module":  row_data.get("module", ""),
                "feature": row_data.get("feature", ""),
                "detail":  row_data.get("detail", ""),
            }

            # 强制：使用知识库分析需求类型（新增/调整）
            from agent.knowledge_manager import get_knowledge_manager
            km = get_knowledge_manager()
            kb_analysis = km.analyze_requirement(req)
            req_type = kb_analysis.get("judgment", "新增")
            
            # 如果是调整需求，获取关联的现有功能作为拆解参考
            related_features = kb_analysis.get("existing_features", [])
            related_modules = kb_analysis.get("related_modules", [])
            
            logger.info(f"[Agent] 需求分析: 行 {row_num}, 类型={req_type}, 关联功能={related_features}")

            # 缓存检查（缓存 key 包含 skill_id 和需求类型）
            cache_req = {**req, "_skill": sid, "_type": req_type}
            cached_result = _get_cached_result(cache_req)
            if cached_result:
                logger.info(f"[Agent] 命中缓存: 行 {row_num}")
                state_out = cached_result
            else:
                code_ctx  = sm.load_code_knowledge(
                    query=f"{req['feature']} {req['detail'][:60]}", limit=2)
                
                # 构建分析上下文
                analysis_context = {
                    "requirement_type": req_type,
                    "related_features": related_features,
                    "related_modules": related_modules,
                }
                
                state_out = graph.invoke({
                    "raw_requirement":  req,
                    "model_id":         model_id,
                    "kb_feature_rules": kb["kb_feature_rules"],
                    "kb_system_caps":   kb["kb_system_caps"],
                    "kb_business_docs": kb["kb_business_docs"],
                    "skill_id":         sid,
                    "skill_config":     skill_config,
                    "skill_examples":   examples,
                    "code_context":     code_ctx,
                    "pages_features":   [],
                    "kb_cases":         [],
                    "g_column_text":    "",
                    "total_days":       0.0,
                    "role_breakdown":   {},
                    "retry_count":      0,
                    "errors":           [],
                    "analysis_context": analysis_context,  # 新增：分析上下文
                })
                _set_cached_result(cache_req, state_out)

            g_text     = state_out.get("g_column_text", "")
            total_days = state_out.get("total_days", 1.0)
            page_count = len(state_out.get("pages_features", []))

            results.append({
                "row":           row_num,
                "g_column_text": g_text,
                "days":          total_days,
                "skipped":       False,
            })

            preview = g_text.split("\n")[0][:60] if g_text else ""
            if progress_callback:
                progress_callback(i + 1, total, row_num, "done", preview, page_count)

        except Exception as e:
            results.append({
                "row":           row_num,
                "g_column_text": f"[处理失败] {str(e)}",
                "days":          None,
                "skipped":       False,
            })
            if progress_callback:
                progress_callback(i + 1, total, row_num, "error", str(e), 0)

        if i < total - 1:
            time.sleep(REQUEST_DELAY)

    return writer.write_results(filepath, results)


def format_eval_result_for_g_column(eval_result: dict) -> str:
    """将智能评估结果格式化为 G 列文本"""
    result = []
    
    # 需求分析
    analysis = eval_result.get("analysis", {})
    result.append(f"【需求分析】")
    result.append(f"类型: {analysis.get('judgment', '未知')}（置信度: {analysis.get('confidence', 0)}%）")
    
    # 初步拆解
    decomposition = eval_result.get("decomposition", [])
    if decomposition:
        result.append("")
        result.append("【初步拆解】")
        for part in decomposition:
            result.append(f"【{part.get('type')}】{part.get('name')}")
            for feat in part.get('features', []):
                result.append(f"  - {feat}")
    
    # 工时评估
    evaluation = eval_result.get("evaluation", {})
    result.append("")
    result.append("【工时评估】")
    result.append(f"模型: {evaluation.get('model', '综合评估')}")
    result.append(f"预估工时: {evaluation.get('effort_days', 0)} 天")
    
    # 建议
    suggestions = analysis.get("suggestions", [])
    if suggestions:
        result.append("")
        result.append("【建议】")
        for suggestion in suggestions:
            result.append(f"• {suggestion}")
    
    return "\n".join(result)


def run_text(text: str, model_id: str = None, progress_callback=None, skill_id: str = None) -> dict:
    """
    处理单条文本需求（不写 Excel）。
    progress_callback: fn(current, total, row_num, status, preview, page_count)
    返回: {"g_text": str, "total_days": float, "page_count": int, "role_breakdown": dict}
    """
    return run_chat(text, model_id=model_id, progress_callback=progress_callback,
                    context="", skill_id=skill_id)


def run_chat(text: str, model_id: str = None, progress_callback=None,
             context: str = "", skill_id: str = None, last_evaluation: dict = None,
             thinking_callback=None, session_knowledge: str = None) -> dict:
    """
    处理聊天消息（支持上下文 + 可切换技能 + 反馈重新评估）。
    - 需求描述过短时返回澄清追问（needs_clarification=True）
    - 通过 skill_id 注入对应技能的 prompt 规则 + 历史案例 + 代码知识库
    - 支持根据反馈重新评估（传入 last_evaluation 参数）
    - 支持 thinking_callback 输出思考状态
    返回:
    {
      "g_text": str,
      "total_days": float,
      "page_count": int,
      "role_breakdown": dict,
      "pages_features": list,
      "is_question": bool,
      "needs_clarification": bool,
      "clarification_question": str,  # 当 needs_clarification=True 时有值
      "is_feedback": bool,            # 是否是反馈重新评估
    }
    """
    from agent import skill_manager as sm

    text = text.strip()
    if not text:
        raise ValueError("消息内容为空")

    # ── 反馈模式（判断是否是对评估结果的反馈，需要重新评估） ──────
    is_feedback = _is_feedback(text)
    if is_feedback and last_evaluation:
        logger.info(f"[Agent-Chat] 反馈重新评估模式: {text[:40]}")
        if thinking_callback:
            thinking_callback("正在分析您的反馈...")
        # 使用历史评估结果作为基础，结合反馈重新评估
        return _re_evaluate_with_feedback(text, context, model_id, skill_id, last_evaluation)

    # ── 追问模式（判断是否是对已有评估结果的追问） ────────────
    is_question = _is_question(text)
    if is_question and context:
        logger.info(f"[Agent-Chat] 追问模式: {text[:40]}")
        if thinking_callback:
            thinking_callback("正在分析您的问题...")
        response = _answer_question(text, context, model_id)
        return {
            "g_text": response, "total_days": 0.0, "page_count": 0,
            "role_breakdown": {}, "pages_features": [],
            "is_question": True, "needs_clarification": False,
            "clarification_question": "",
            "is_feedback": False,
        }

    # ── 拆解需求描述 ─────────────────────────────────────────
    lines        = text.split('\n')
    feature      = lines[0].strip()[:80]
    detail_lines = [l for l in lines[1:] if l.strip()]
    detail       = '\n'.join(detail_lines) if detail_lines else text

    # 合并临时知识库和历史对话上下文
    if session_knowledge and context:
        detail = f"【用户上传的参考文档】\n{session_knowledge}\n\n【历史对话】\n{context}\n\n【当前需求】\n{detail}"
    elif session_knowledge:
        detail = f"【用户上传的参考文档】\n{session_knowledge}\n\n【当前需求】\n{detail}"
    elif context:
        detail = f"【历史对话】\n{context}\n\n【当前需求】\n{detail}"

    # ── 澄清检查：描述过于简短时返回追问，避免无效评估 ────────
    raw_detail = '\n'.join(detail_lines) if detail_lines else text
    if len(raw_detail.strip()) < 15 and len(feature.strip()) < 15:
        question = f"您提到要做「{feature or text}」，能详细说说具体要实现哪些功能，涉及哪些页面或操作流程吗？"
        return {
            "g_text": "", "total_days": 0.0, "page_count": 0,
            "role_breakdown": {}, "pages_features": [],
            "is_question": False, "needs_clarification": True,
            "clarification_question": question,
        }

    req = {
        "row": 1, "module": "", "feature": feature,
        "detail": detail, "extra": "",
        "existing_g": None, "existing_n": None,
    }

    # ── 加载技能配置 + 历史案例 + 代码知识库 ─────────────────
    if thinking_callback:
        thinking_callback("正在加载知识库和技能配置...")
    
    sid          = skill_id or sm.get_current_skill_id()
    skill_config = sm.get_skill(sid)
    examples     = sm.load_examples(sid, limit=3)
    code_context = sm.load_code_knowledge(query=f"{feature} {raw_detail[:60]}", limit=2)

    logger.info(f"[Agent-Chat] 需求={feature[:40]} 技能={sid} 历史案例={len(examples)} 代码知识={bool(code_context)}")

    # ── 加载知识库 ─────────────────────────────────────────
    if thinking_callback:
        thinking_callback("正在加载企业知识库...")
    
    kb    = KnowledgeLoader().load()
    
    # ── 构建并执行工作流 ───────────────────────────────────
    if thinking_callback:
        thinking_callback("正在构建分析流程...")
    
    graph = build_graph()

    if thinking_callback:
        thinking_callback("正在识别需求类型...")

    if thinking_callback:
        thinking_callback("正在进行结构化拆解...")

    if thinking_callback:
        thinking_callback("正在核算工时评估...")

    state = graph.invoke({
        "raw_requirement":  req,
        "model_id":         model_id,
        "kb_feature_rules": kb["kb_feature_rules"],
        "kb_system_caps":   kb["kb_system_caps"],
        "kb_business_docs": kb["kb_business_docs"],
        # 技能注入
        "skill_id":       sid,
        "skill_config":   skill_config,
        "skill_examples": examples,
        "code_context":   code_context,
        # 初始化输出字段
        "pages_features":  [],
        "kb_cases":        [],
        "g_column_text":   "",
        "total_days":      0.0,
        "role_breakdown":  {},
        "retry_count":     0,
        "errors":          [],
    })

    # ── 整理结果 ─────────────────────────────────────────
    if thinking_callback:
        thinking_callback("正在整理评估结果...")

    g_text         = state.get("g_column_text", "")
    total_days     = state.get("total_days", 1.0)
    role_breakdown = state.get("role_breakdown", {})
    pages_features = state.get("pages_features", [])
    page_count     = len(pages_features)

    if progress_callback:
        preview = g_text.split("\n")[0][:60] if g_text else ""
        progress_callback(1, 1, 1, "done", preview, page_count)

    if thinking_callback:
        thinking_callback("评估完成！")

    return {
        "g_text": g_text, "total_days": total_days, "page_count": page_count,
        "role_breakdown": role_breakdown, "pages_features": pages_features,
        "is_question": False, "needs_clarification": False,
        "clarification_question": "",
        "is_feedback": False,
    }


def _re_evaluate_with_feedback(feedback: str, context: str, model_id: str, skill_id: str, last_evaluation: dict) -> dict:
    """
    根据用户反馈重新评估需求
    :param feedback: 用户反馈内容（原始用户消息，非 full_text）
    :param context: 历史对话上下文（含原始评估结果）
    :param model_id: 模型ID
    :param skill_id: 技能ID
    :param last_evaluation: 上一次评估结果（可选）
    :return: 重新评估结果
    """
    from agent import gemini_client

    original_eval_text = ""
    original_summary = ""
    if last_evaluation:
        original_eval_text = last_evaluation.get('g_text', '')[:800]
        total_days = last_evaluation.get('total_days', 0)
        role_breakdown = last_evaluation.get('role_breakdown', {})
        role_str = "、".join(f"{k}:{v}天" for k, v in role_breakdown.items()) if role_breakdown else "未知"
        original_summary = f"总工时：{total_days}天，角色分配：{role_str}"

    prompt = f"""你是一位有10年经验的IT项目经理和产品专家。

## 对话历史（含原始需求和上次评估结果）
{context}

## 上次评估摘要
{original_summary}

## 上次评估详情
{original_eval_text}

## 用户的修改反馈
{feedback}

## 任务
根据用户反馈，对上次的工时评估进行调整，输出修正后的完整拆解和工时。

## 输出格式（严格JSON）
{{
  "g_text": "修正后的完整拆解说明（包含各页面/功能点和工时）",
  "total_days": 5.0,
  "role_breakdown": {{"产品/设计": 1.0, "前端开发": 1.0, "后端开发": 2.0, "测试": 1.0}},
  "reason": "调整原因说明"
}}

## 注意
- g_text 要包含完整的拆解内容，不能只说"已调整"
- total_days 精确到0.5天
- 只输出JSON，不要其他内容"""

    try:
        response = gemini_client.call_llm(prompt, model_id=model_id)
        import json
        text = response.strip().replace('```json', '').replace('```', '')
        data = json.loads(text)
        
        return {
            "g_text": data.get("g_text", response),
            "total_days": data.get("total_days", 1.0),
            "page_count": 0,
            "role_breakdown": data.get("role_breakdown", {}),
            "pages_features": [],
            "is_question": False,
            "needs_clarification": False,
            "clarification_question": "",
            "is_feedback": True,
        }
    except Exception as e:
        logger.error(f"重新评估失败: {e}")
        # 降级处理：基于反馈简单调整
        adjusted_days = _adjust_worktime_by_feedback(last_evaluation, feedback)
        return {
            "g_text": f"【反馈调整】{feedback}\n\n根据您的反馈，重新评估结果如下：\n{adjusted_days.get('g_text', '')}",
            "total_days": adjusted_days.get("total_days", 1.0),
            "page_count": 0,
            "role_breakdown": adjusted_days.get("role_breakdown", {}),
            "pages_features": [],
            "is_question": False,
            "needs_clarification": False,
            "clarification_question": "",
            "is_feedback": True,
        }


def _adjust_worktime_by_feedback(last_evaluation: dict, feedback: str) -> dict:
    """
    基于用户反馈简单调整工时（降级策略）
    """
    import re
    
    result = {
        "total_days": last_evaluation.get("total_days", 1.0),
        "role_breakdown": last_evaluation.get("role_breakdown", {}),
        "g_text": "",
    }
    
    # 根据关键词调整工时
    if "太高" in feedback or "调低" in feedback or "减少" in feedback:
        # 降低20%-30%
        factor = 0.75 if "大幅" in feedback else 0.85
        result["total_days"] = round(result["total_days"] * factor * 2) / 2
        result["g_text"] = f"根据反馈，工时下调约{int((1-factor)*100)}%，调整为 {result['total_days']} 天"
    
    elif "太低" in feedback or "调高" in feedback or "增加" in feedback:
        # 增加15%-30%
        factor = 1.3 if "大幅" in feedback else 1.15
        result["total_days"] = round(result["total_days"] * factor * 2) / 2
        result["g_text"] = f"根据反馈，工时上调约{int((factor-1)*100)}%，调整为 {result['total_days']} 天"
    
    elif "重新评估" in feedback or "重新计算" in feedback:
        result["g_text"] = "已重新评估，结果保持不变"
    
    # 调整特定角色工时
    role_pattern = r'(前端|后端|产品|测试).*(\d+\.?\d*)天'
    match = re.search(role_pattern, feedback)
    if match:
        role = match.group(1)
        days = float(match.group(2))
        
        # 映射角色名称
        role_mapping = {
            "前端": "前端开发",
            "后端": "后端开发", 
            "产品": "产品/设计",
            "测试": "测试",
        }
        role_key = role_mapping.get(role, role)
        
        result["role_breakdown"][role_key] = days
        # 重新计算总工时
        result["total_days"] = round(sum(result["role_breakdown"].values()) * 2) / 2
        result["g_text"] = f"已调整{role}工时为 {days} 天，总工时更新为 {result['total_days']} 天"
    
    return result


def _is_question(text: str) -> bool:
    """判断是否是追问/问题（非需求拆解请求）"""
    question_patterns = [
        r'^[你您][觉认]得.*',
        r'^为什么.*',
        r'^什么.*',
        r'^怎么.*',
        r'^如何.*',
        r'^能不能.*',
        r'^可以.*吗\??$',
        r'^是否.*',
        r'^应该.*',
        r'^如果.*',
        r'.*怎么样\??$',
        r'.*好不好\??$',
        r'.*对吗\??$',
        r'.*是吗\??$',
        r'.*呢\??$',
        r'.*吧\??$',
        r'^再.*',
        r'^继续.*',
        r'^更多.*',
        r'^详细.*',
        r'^解释.*',
        r'^说明.*',
        r'^分析.*',
        r'^比较.*',
    ]
    
    import re
    text = text.strip()
    for pattern in question_patterns:
        if re.match(pattern, text):
            return True
    return False


def _is_feedback(text: str) -> bool:
    """判断是否是评估反馈/修正请求（需要重新评估）"""
    feedback_patterns = [
        # ========== 数值调整类 ==========
        r'.*太高.*',
        r'.*太低.*',
        r'.*偏高.*',
        r'.*偏低.*',
        r'.*调整.*工时',
        r'.*工时.*调整',
        r'.*修改.*工时',
        r'.*工时.*修改',
        r'.*改.*工时',
        r'.*工时.*改',
        r'.*调低.*',
        r'.*调高.*',
        r'.*减少.*天',
        r'.*增加.*天',
        r'.*加.*天',
        r'.*减.*天',
        r'.*天数.*不对',
        r'.*工日.*调整',
        
        # ========== 重新评估类 ==========
        r'重新评估',
        r'重新计算',
        r'重新拆解',
        r'再算一遍',
        r'再评估一次',
        r'重新算一下',
        r'重新估一下',
        r'帮我再算',
        r'重新来',
        r'重新分析',
        r'重新评估一下',
        r'重新计算一下',
        
        # ========== 否定/质疑类 ==========
        r'.*不合理.*',
        r'.*不准确.*',
        r'.*不对.*',
        r'.*错了.*',
        r'.*有问题.*',
        r'.*估时.*不准',
        r'.*估算.*不准',
        r'.*评估.*不准',
        r'.*工时.*不准',
        r'.*天数.*不准',
        r'.*算错.*',
        r'.*少算了.*',
        r'.*多算了.*',
        r'.*漏算了.*',
        
        # ========== 功能调整类 ==========
        r'.*漏了.*功能',
        r'.*多了.*功能',
        r'.*缺少.*功能',
        r'.*不需要.*',
        r'.*增加.*功能',
        r'.*去掉.*功能',
        r'.*移除.*功能',
        r'.*补充.*功能',
        r'.*遗漏.*功能',
        
        # ========== 确认/指示类 ==========
        r'按这个来',
        r'按这个调整',
        r'按我说的算',
        r'按这个修改',
        r'按照这个来',
        
        # ========== 疑问/请求类 ==========
        r'.*能不能.*再算',
        r'.*能不能.*重新',
        r'.*可以.*重新评估',
        r'.*应该.*多少天',
        r'.*需要.*多少天',
        r'.*工时.*应该是',
        r'.*天数.*应该是',
    ]
    
    import re
    text = text.strip()
    for pattern in feedback_patterns:
        if re.search(pattern, text):
            return True
    return False


def _answer_question(question: str, context: str, model_id: str = None) -> str:
    """回答追问问题"""
    from agent import gemini_client
    
    prompt = f"""你是一位有10年经验的IT项目经理和产品专家。

## 对话历史
{context}

## 当前问题
{question}

## 任务
基于对话历史，回答用户的问题。
如果是关于工时评估的问题，请给出专业建议；
如果是一般性问题，请友好解答。

## 输出要求
- 回答要简明扼要
- 使用自然语言
- 保持专业但友好的语气"""
    
    try:
        response = gemini_client.call_llm(prompt, model_id=model_id)
        return response.strip()
    except Exception as e:
        logger.error(f"回答问题失败: {e}")
        return f"抱歉，我暂时无法回答这个问题。错误: {str(e)}"



def export_to_excel(results: Dict[str, Any], output_path: str = None) -> str:
    """
    导出评估结果到Excel
    :param results: 评估结果
    :param output_path: 输出路径，不传则自动生成
    :return: 生成的文件路径
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = f"evaluation_result_{timestamp}.xlsx"
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "需求拆解评估"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 表头（新增接口列）
    headers = ["序号", "原始需求", "需求类型", "已拆解需求", "接口列表", "产品工时(天)", "评估模型", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # 填充数据
    row_num = 2
    for i, result in enumerate(results["results"], 1):
        req = result["original_requirement"]
        analysis = result["analysis"]
        evaluation = result["evaluation"]

        # 序号
        ws.cell(row=row_num, column=1, value=i).border = thin_border

        # 原始需求
        original_req = f"{req.get('module', '')} - {req.get('feature', '')}\n{req.get('detail', '')}"
        ws.cell(row=row_num, column=2, value=original_req).border = thin_border

        # 需求类型（新增/调整）
        ws.cell(row=row_num, column=3, value=analysis.get("judgment", "未知")).border = thin_border

        # 已拆解需求
        decomposition = result["decomposition"]
        decomposed_text = ""
        for part in decomposition:
            decomposed_text += f"【{part.get('type')}】{part.get('name')}\n"
            for feat in part.get('features', []):
                decomposed_text += f"  - {feat}\n"
        ws.cell(row=row_num, column=4, value=decomposed_text.strip()).border = thin_border

        # 接口列表（新增字段）
        interfaces = []
        if isinstance(decomposition, list):
            for part in decomposition:
                if "interfaces" in part and part["interfaces"]:
                    interfaces.extend(part["interfaces"])
        interfaces_text = "、".join(interfaces) if interfaces else "新增接口"
        ws.cell(row=row_num, column=5, value=interfaces_text).border = thin_border

        # 产品工时
        ws.cell(row=row_num, column=6, value=evaluation.get("effort_days", 0)).border = thin_border

        # 评估模型
        ws.cell(row=row_num, column=7, value=evaluation.get("model", "综合评估")).border = thin_border

        # 备注
        related_modules = analysis.get("related_modules", [])
        suggestions = analysis.get("suggestions", [])
        remark = ""
        if related_modules:
            remark += f"相关模块: {', '.join(related_modules)}\n"
        if suggestions:
            remark += "\n".join(suggestions)
        ws.cell(row=row_num, column=8, value=remark.strip()).border = thin_border

        row_num += 1

    # 添加合计行
    total_row = row_num
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1, value="合计").border = thin_border
    ws.cell(row=total_row, column=6, value=results["total_days"]).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=results["total_days"]).border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # 设置行高
    for row in range(1, total_row + 1):
        ws.row_dimensions[row].height = 15
    
    # 保存文件
    wb.save(output_path)
    logger.info(f"评估结果已导出到: {output_path}")

    return output_path


# ============================================================
# Step 8: 表格格式化和接口识别函数
# ============================================================

def format_evaluation_as_table(evaluation_result: dict, session_knowledge: str = None) -> str:
    """
    将评估结果格式化为表格格式

    :param evaluation_result: 评估结果字典，包含 pages_features, total_days, role_breakdown 等
    :param session_knowledge: 用户上传的会话知识库（用于接口识别）
    :return: 格式化的表格字符串
    """
    pages_features = evaluation_result.get("pages_features", [])
    total_days = evaluation_result.get("total_days", 0)
    role_breakdown = evaluation_result.get("role_breakdown", {})

    if not pages_features:
        return "未生成评估结果"

    # 构建表格头
    lines = []
    lines.append("| 序号 | 改造点 | 页面类型 | 功能描述 | 接口 | 工时(天) |")
    lines.append("|------|--------|---------|---------|------|----------|")

    # 添加数据行
    for idx, item in enumerate(pages_features, 1):
        page_name = item.get("页面", "")
        page_type = item.get("类型", "新增")
        features = item.get("功能点", [])
        interfaces = item.get("接口", [])
        effort = item.get("工时", 0)

        # 功能描述（多个功能点用、分隔）
        features_text = "、".join(features) if features else "（无）"

        # 接口信息
        if interfaces:
            interfaces_text = "、".join(interfaces) if isinstance(interfaces, list) else interfaces
        else:
            interface_count = len(features)
            interfaces_text = f"新增{interface_count}个" if interface_count > 0 else "（无）"

        # 页面信息
        page_info = f"{page_name}({page_type})"

        # 添加行
        lines.append(f"| {idx} | {page_info} | {page_type} | {features_text} | {interfaces_text} | {effort} |")

    # 添加合计行
    lines.append("|------|--------|---------|---------|------|----------|")
    role_text = "、".join(f"{k}:{v}天" for k, v in role_breakdown.items()) if role_breakdown else "未评估"
    lines.append(f"| 合计 | - | - | - | - | **{total_days}天** |")
    lines.append(f"\n工时分配：{role_text}")

    return "\n".join(lines)


def extract_interfaces_from_decomposition(decomposition_item: dict) -> list:
    """
    从拆解点中提取接口信息

    :param decomposition_item: 单个拆解点 {"页面": "...", "功能点": [...], "类型": "..."}
    :return: 提取的接口列表
    """
    import re

    interfaces = []

    # 获取功能点列表
    features = decomposition_item.get("功能点", [])

    # 从功能点描述中用正则提取接口名
    patterns = [
        r'(?:接口|API|api)[\s：:]*([a-zA-Z0-9_\-/]+)',
        r'(?:调用|请求)[\s：:]*([a-zA-Z0-9_\-/]+)',
    ]

    for feature in features:
        for pattern in patterns:
            matches = re.findall(pattern, str(feature))
            interfaces.extend(matches)

    return list(set(interfaces)) if interfaces else []


def search_interfaces_in_docs(decomposition_title: str, session_knowledge: str) -> list:
    """
    在用户上传的文档中搜索与分解点相关的接口

    :param decomposition_title: 拆解点标题
    :param session_knowledge: 用户上传的会话知识库内容
    :return: 匹配到的接口列表
    """
    import re

    if not session_knowledge:
        return []

    interfaces = set()

    # 接口名称识别模式（多种格式）
    patterns = [
        r'(?:接口|API|api|endpoint)[\s：:]*([a-zA-Z0-9_\.\-/]+)',
        r'(?:方法|Method|GET|POST|PUT|DELETE)[\s：:]*([a-zA-Z0-9_\-/]+)',
        r'def\s+([a-zA-Z0-9_]+)\s*\(',
        r'function\s+([a-zA-Z0-9_]+)\s*\(',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, session_knowledge)
        interfaces.update(matches)

    # 返回前20个唯一接口
    return list(interfaces)[:20]


